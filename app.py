# =====================================
# Streamlit Web App: 不担保人事用合同记录表自动审核（改进版）
# =====================================

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from datetime import datetime

st.title("不担保人事用合同记录表自动审核")

# -------- 上传文件 ----------
uploaded_files = st.file_uploader(
    "请上传以下文件：不担保表, 放款明细, 字段, 二次明细, 重卡数据",
    type="xlsx",
    accept_multiple_files=True
)

if not uploaded_files or len(uploaded_files) < 5:
    st.warning("请上传所有 5 个文件")
    st.stop()
else:
    st.success("✅ 文件上传完成")

# -------- 工具函数 ----------
def find_file(files_list, keyword):
    for f in files_list:
        if keyword in f.name:
            return f
    raise FileNotFoundError(f"❌ 未找到包含关键词「{keyword}」的文件")

def normalize_colname(c):
    return str(c).strip().lower()

def find_col(df, keyword):
    if df is None:
        return None
    key = keyword.strip().lower()
    for col in df.columns:
        if key in normalize_colname(col):
            return col
    return None

def find_sheet(xls, keyword):
    for s in xls.sheet_names:
        if keyword in s:
            return s
    raise ValueError(f"❌ 未找到包含关键词「{keyword}」的sheet")

# -------- 读取文件 ----------
main_file = find_file(uploaded_files, "不担保")
fk_file   = find_file(uploaded_files, "放款明细")
zd_file   = find_file(uploaded_files, "字段")
ec_file   = find_file(uploaded_files, "二次明细")
zk_file   = find_file(uploaded_files, "重卡数据")

# 主表 sheet 模糊匹配 "二次"
xls_main = pd.ExcelFile(main_file)
target_sheet = find_sheet(xls_main, "二次")
main_df = pd.read_excel(xls_main, sheet_name=target_sheet, header=1)  # 第二行为列名

# 放款明细 sheet 模糊匹配 "本司"
xls_fk = pd.ExcelFile(fk_file)
fk_sheet = find_sheet(xls_fk, "本司")
fk_df = pd.read_excel(xls_fk, sheet_name=fk_sheet, header=0)

# 字段 sheet 模糊匹配 "重卡"
xls_zd = pd.ExcelFile(zd_file)
zd_sheet = find_sheet(xls_zd, "重卡")
zd_df = pd.read_excel(xls_zd, sheet_name=zd_sheet, header=0)

# 二次明细和重卡数据
ec_df = pd.read_excel(ec_file, header=0)
zk_df = pd.read_excel(zk_file, header=0)

# -------- 字段映射 ----------
mapping_fk = {
    "授信方": "授信",
    "租赁本金": "本金",
    "租赁期限": "期限",
    "客户经理": "客户经理",
    "起租收益率": "收益率",
    "主车台数": "主车台数",
    "挂车台数": "挂车台数"
}
mapping_zd = {
    "保证金比例": "保证金比例_2",
    "项目提报人": "提报",
    "起租时间": "起租日_商",
    "租赁期限": "总期数_商_资产"
}
mapping_ec = {"二次时间": "出本流程时间_节点"}
mapping_zk = {"结清日期": "核销"}

# -------- 输出准备 ----------
output_path = "不担保人事用合同记录表_审核标注版.xlsx"
empty_row = pd.DataFrame([[""] * len(main_df.columns)], columns=main_df.columns)
main_df_with_blank = pd.concat([empty_row, main_df], ignore_index=True)
main_df_with_blank.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# -------- 找合同列 ----------
contract_col_main = find_col(main_df, "合同")
contract_col_fk = find_col(fk_df, "合同")
contract_col_zd = find_col(zd_df, "合同")
contract_col_ec = find_col(ec_df, "合同")
contract_col_zk = find_col(zk_df, "合同")

if not contract_col_main:
    st.error("❌ 在主表中未能找到包含关键词 '合同' 的列，请确认列名。")
    st.stop()

# -------- 比对函数 ----------
def normalize_num(val):
    """去除百分号并尝试转为浮点数"""
    if pd.isna(val):
        return None
    s = str(val).replace("%", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return s  # 若不是数值，返回原文本

def same_date_ymd(a, b):
    """判断两个日期是否年月日一致"""
    try:
        da = pd.to_datetime(a, errors='coerce')
        db = pd.to_datetime(b, errors='coerce')
        if pd.isna(da) or pd.isna(db):
            return False
        return (da.year == db.year) and (da.month == db.month) and (da.day == db.day)
    except Exception:
        return False

def compare_fields_and_mark(row_idx, row, main_df, main_kw, ref_df, ref_kw, ref_contract_col):
    errors = 0
    main_col = find_col(main_df, main_kw)
    ref_col = find_col(ref_df, ref_kw)
    if not main_col or not ref_col or not ref_contract_col:
        return 0
    contract_no = str(row.get(contract_col_main)).strip()
    if pd.isna(contract_no) or contract_no == "nan" or contract_no == "":
        return 0
    ref_rows = ref_df[ref_df[ref_contract_col].astype(str).str.strip() == contract_no]
    if ref_rows.empty:
        return 0
    ref_val = ref_rows.iloc[0][ref_col]
    main_val = row.get(main_col)

    if pd.isna(main_val) and pd.isna(ref_val):
        return 0

    # ---- 日期字段模糊比较 ----
    if "时间" in main_kw or "日期" in main_kw or "时间" in ref_kw or "日期" in ref_kw:
        if not same_date_ymd(main_val, ref_val):
            errors = 1
    else:
        # ---- 数值与文本混合比较 ----
        main_num = normalize_num(main_val)
        ref_num = normalize_num(ref_val)

        if isinstance(main_num, (int, float)) and isinstance(ref_num, (int, float)):
            if abs(main_num - ref_num) > 1e-6:
                errors = 1
        else:
            if str(main_num).strip() != str(ref_num).strip():
                errors = 1

    # ---- 标红 ----
    if errors:
        excel_row = row_idx + 3  # header=1 + 空行
        col_idx = list(main_df.columns).index(main_col) + 1
        ws.cell(excel_row, col_idx).fill = red_fill
    return errors

# -------- 主循环 ----------
total_errors = 0
for idx, row in main_df.iterrows():
    if pd.isna(row.get(contract_col_main)):
        continue
    for main_kw, ref_kw in mapping_fk.items():
        total_errors += compare_fields_and_mark(idx, row, main_df, main_kw, fk_df, ref_kw, contract_col_fk)
    for main_kw, ref_kw in mapping_zd.items():
        total_errors += compare_fields_and_mark(idx, row, main_df, main_kw, zd_df, ref_kw, contract_col_zd)
    for main_kw, ref_kw in mapping_ec.items():
        total_errors += compare_fields_and_mark(idx, row, main_df, main_kw, ec_df, ref_kw, contract_col_ec)
    for main_kw, ref_kw in mapping_zk.items():
        total_errors += compare_fields_and_mark(idx, row, main_df, main_kw, zk_df, ref_kw, contract_col_zk)

# -------- 黄色标记合同号 ----------
contract_col_idx_excel = list(main_df.columns).index(contract_col_main) + 1
for row_idx in range(len(main_df)):
    excel_row = row_idx + 3
    has_red = any(ws.cell(excel_row, c).fill == red_fill for c in range(1, len(main_df.columns) + 1))
    if has_red:
        ws.cell(excel_row, contract_col_idx_excel).fill = yellow_fill

# -------- 保存并提供下载 ----------
output = BytesIO()
wb.save(output)
output.seek(0)

st.success(f"✅ 审核完成，共发现 {total_errors} 处不一致。")
st.download_button(
    label="下载审核标注版 Excel",
    data=output,
    file_name="不担保人事用合同记录表_审核标注版.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
